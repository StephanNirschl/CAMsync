import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from config.config_loader import ConfigLoader
import re


def sanitize_sheet_title(name: str) -> str:
    """
    Bereinigt den Namen für ein Excel-Tabellenblatt:
    Erlaubt keine Sonderzeichen wie / \ * ? [ ] etc.
    """
    name = re.sub(r'[\\/*?:[\]]', '_', name)
    name = name.replace('|', '_').replace('\n', ' ')
    return name[:31]


def export_all_status_tools_to_excel(config_loader: ConfigLoader, logger):
    try:
        db_path = config_loader.get("tools", "db")
        if not os.path.exists(db_path):
            logger(f"❌ Datenbankpfad nicht gefunden: {db_path}")
            return

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Werkzeugstatus-ID aus CustomDataClasses ermitteln
        cursor.execute("""
            SELECT custom_data_class_id
            FROM CustomDataClasses
            WHERE name = 'Werkzeugstatus'
            """)
        row = cursor.fetchone()
        if not row:
            logger("❌ 'Werkzeugstatus' nicht in CustomDataClasses gefunden.")
            return
        werkzeugstatus_class_id = row[0]

        cursor.execute("""
            SELECT DISTINCT custom_data_value
            FROM CustomDataValues
            WHERE custom_data_class_id = ?
        """, (werkzeugstatus_class_id,))


        status_list = [row[0] for row in cursor.fetchall()]
        if not status_list:
            logger("ℹ️ Keine Werkzeugstatus mit custom_data_class_id = 3 gefunden.")
            return

        wb = Workbook()
        wb.remove(wb.active)
        status_counts = {}

        for status in status_list:
            logger(f"📦 Exportiere Status: {status}")

            query = """
            SELECT
                g.name AS klasse,
                t.nc_number_val AS id,
                t.nc_name AS name,
                t.comment AS kommentar,
                tt.dbl_param4 AS werkzeug_durchmesser,
                tt.dbl_param5 AS spitzenlänge,
                t.gage_length AS gesamtlänge,
                v.custom_data_value AS status,
                c.class AS coupling_klasse
            FROM
                NCTools t
            JOIN
                NCToolCustomData cd ON t.id = cd.nctool_id
            JOIN
                CustomDataValues v ON cd.custom_data_value_id = v.custom_data_value_id
            LEFT JOIN
                Tools tt ON t.tool_id = tt.id
            LEFT JOIN
                GeometryClasses g ON tt.tool_type_id = g.id
            LEFT JOIN
                Holders h ON t.holder_id = h.id
            LEFT JOIN
                Couplings c ON h.top_coupling_id = c.coupling_id
            WHERE
                v.custom_data_value = ?
            """

            cursor.execute(query, (status,))
            results = cursor.fetchall()
            status_counts[status] = len(results)

            if not results:
                logger(f"⚠️ Keine Daten für Status '{status}' gefunden.")
                continue

            sheet_title = sanitize_sheet_title(status)
            ws = wb.create_sheet(title=sheet_title)
            ws.append([
                "Klasse", "ID", "NAME", "Kommentar", 
                "Wkz-Ø", "S-L", "G-L", "Status", "Kupplungs-Klasse"
            ])

            for row in results:
                ws.append(row)

            end_row = ws.max_row
            end_col = ws.max_column
            table_range = f"A1:{chr(64 + end_col)}{end_row}"

            table = Table(displayName=f"TBL_{status[:20].replace(' ', '_')}", ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)

            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2

        # Übersicht
        summary_ws = wb.create_sheet(title="Übersicht", index=0)
        summary_ws.append(["Status", "Anzahl Werkzeuge", "Anteil (%)"])

        total = sum(status_counts.values())
        for status, count in status_counts.items():
            percentage = round((count / total) * 100, 2)
            summary_ws.append([status, count, f"{percentage:.2f} %"])
        summary_ws.append(["Gesamt", total, "100 %"])

        last_row = summary_ws.max_row
        bold_font = Font(bold=True)
        fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for cell in summary_ws[f"A{last_row}:C{last_row}"][0]:
            cell.font = bold_font
            cell.fill = fill

        for col in summary_ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            summary_ws.column_dimensions[col_letter].width = max_length + 2

        chart = PieChart()
        chart.title = "Verteilung Werkzeugstatus"
        chart.height = 15
        chart.width = 15
        data_range = Reference(summary_ws, min_col=2, min_row=2, max_row=last_row - 1)
        labels_range = Reference(summary_ws, min_col=1, min_row=2, max_row=last_row - 1)
        chart.add_data(data_range, titles_from_data=False)
        chart.set_categories(labels_range)
        summary_ws.add_chart(chart, "A17")

        conn.close()
        db_dir = os.path.dirname(db_path)
        date_str = datetime.now().strftime("%Y%m%d")
        file_name = f"ToolDatabase_report_ALL_{date_str}.xlsx"
        file_path = os.path.join(db_dir, file_name)
        wb.save(file_path)
        logger(f"✅ Mehrblatt-Excel gespeichert: {file_path}")
        os.startfile(file_path)

    except Exception as e:
        logger(f"⚠️ Fehler beim Multi-Export: {e}")
